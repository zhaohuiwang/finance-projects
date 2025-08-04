# -*- coding: utf-8 -*-
"""
Created on Tue Nov 17 14:11:55 2020
@author: Family_Optiplex9
"""

# =============================================================================
# import pandas as pd
# import xlsxwriter
# from datetime import datetime
# 
# export_file = f"finance_{datetime.now().strftime('%m_%d_%Y')}"
# 
# ######## create a workbook and write DataFrames as new worksheets ########
# 
# import xlsxwriter
# # create a Pandas Excel writer using xlsxwriter as a engine
# writer = pd.ExcelWriter(f'C:\\Users\\Family_Optiplex9\\Downloads\\{export_file}.xlsx',
#                         engine='xlsxwriter')
# # write each dataframe to a different worksheet
# df.to_excel(writer, sheet_name='sheet1_name')
# # closed the Pandas Excel writer and output the Excel file
# writer.save()
# 
# 
# ##### open an existing workbood and append DataFrames as new worksheets ######
# 
# # mode='a' the workbook must exist, it will open it and append more worksheet 
# # mode='w' create, overwrite if one exist, open it and write worksheets
# # one could just use engine='openpyxl', mode='w' to create a new workbook, and 
# # engine='openpyxl', mode='a' to append more worksheet, and 
# # no need for the above engine='xlsxwriter' 
# 
# 
# from openpyxl import load_workbook
# # create a Pandas Excel writer using openpyxl as a engine
# writer = pd.ExcelWriter(f'C:\\Users\\Family_Optiplex9\\Downloads\\{export_file}.xlsx',
#                         engine='openpyxl', mode='a')
# df.to_excel(writer, sheet_name='sheet2_name')
# # closed the Pandas Excel writer and output the Excel file
# writer.save()

# # readin all worksheets from a workbook as DataFrame dictionary
# # method 1
# from datetime import datetime 
# export_file = f"finance_{datetime.now().strftime('%m_%d_%Y')}"

# # conda install xlrd
# import xlrd
# xlsx = pd.ExcelFile(f'C:\\Users\\Family_Optiplex9\\Downloads\\{export_file}.xlsx')
# df = pd.read_excel(xlsx, sheet_name=None)
# # sheet_name can be a name, list of number or names
# # if None is specified, all sheets are returned,
# # as a {sheet_name:dataframe} dictionary

# # method 2
# sheet_to_df_map = {}
# for sheet_name in xlsx.sheet_names:
#     sheet_to_df_map[sheet_name] = xlsx.parse(sheet_name)
    
# =============================================================================



import numpy as np
import requests
import pandas as pd
import re
import os
import xlsxwriter
from datetime import datetime
# os.chdir('C:\\Users\\Family_Optiplex9\\Downloads')
export_file = f"finance_{datetime.now().strftime('%m_%d_%Y')}"

def str_f(pd_series):
    '''
    convert a pandas.Series with values like '123.123M' to 123.123

    Parameters
    ----------
    pd_series : pandas.Series

    Returns: pandas.Series of the same length as the input
    -------
    None.

    '''
    return pd_series.str.strip('Mm%').to_numpy(dtype='float')

url_mw = 'https://www.marketwatch.com/tools/ipo-calendar'
url_nyse = 'https://www.nyse.com/ipo-center/recent-ipo' 
url_nasdaq = 'https://www.nasdaq.com/market-activity/ipos'

url_iposcoop = 'https://www.iposcoop.com/current-year-pricings/'

html_iposcoop = requests.get(url_iposcoop).content
# Pandas.read_html(): Read HTML tables into a list of DataFrame objects.
df_list_iposcoop = pd.read_html(html_iposcoop)
df_iposcoop = df_list_iposcoop[0]
df_iposcoop.columns
df_iposcoop['Return'] = str_f(df_iposcoop['Return']).round()

df = df_iposcoop.sort_values(
    ['Industry', 'Return'], axis=0, ascending=False, inplace=False,
               kind='quicksort', na_position='last',
               ignore_index=True, key=None)

df.columns = df.columns.str.replace(r'%', 'per') # '%' to 'per'
df.columns = df.columns.str.replace(r'\s+', '_', regex=True)  # whitespaces to '_'
df.columns = df.columns.str.strip('_ ')          # remove leading/trailing '_'
df.columns = map(str.lower, df.columns)          # to all lower case


# print(df)
# df.to_csv('C:\\Users\\Family_Optiplex9\\Downloads\\my_data.csv')

# groupby X coumn and then sort within each group
df1 = df.groupby('industry', group_keys=False).apply(
    lambda x: x.sort_values('return', ascending=False))

# groupby X coumn and then sort within each group
df2 = df.groupby('industry', group_keys=False).apply(
    lambda x: x.sort_values('offer_date', ascending=False))


######## create a workbook and write DataFrames as new worksheets ########
import pandas as pd
import xlsxwriter

# create a Pandas Excel writer using xlsxwriter as a engine
writer = pd.ExcelWriter(f'C:\\Users\\Family_Optiplex9\\Downloads\\{export_file}.xlsx',
                        engine='xlsxwriter')

# write each dataframe to a different worksheet
df1.to_excel(writer, sheet_name='ipo_by_return')
df2.to_excel(writer, sheet_name='ipo_by_offer_date')

# closed the Pandas Excel writer and output the Excel file
writer.save()


# =============================================================================
# # to display everythin in the dataframe, all columns for a row in single line
# pd.set_option('display.max_columns', None, 
#                   'display.max_rows', None, 
#                   'display.width', 1000, 
#                   'display.max_colwidth', 0)
# # reset the display to the default
# pd.reset_option('all')
# 
# =============================================================================



'''
Yahoo most active stock screener
'''
n_per_page = 25
new_set = 0 
df = pd.DataFrame()
while new_set >= 0:
    url =f'https://finance.yahoo.com/most-active/?count={n_per_page}&offset={new_set}'  
    # Pandas.read_html(): Read HTML tables into a list of DataFrame objects.
    new_df_list = pd.read_html(url)
    new_df = new_df_list[0]
    if len(new_df) == 0:
        break
    df = pd.concat([df, new_df], axis=0, ignore_index=True)
    new_set += (n_per_page + 1) 
# df.columns
# df.head()

df.columns = df.columns.str.replace(r'%', 'per') # '%' to 'per'
df.columns = df.columns.str.replace('\W+', '_')  # whitespaces to '_'
df.columns = df.columns.str.strip('_ ')          # remove leading/trailing '_'
df.columns = map(str.lower, df.columns)          # to all lower case 

# change string values to float64 for sorting
df['volume_ratio'] = str_f(df['volume']) / str_f(df['avg_vol_3_month'])   
df['per_change'] = str_f(df['per_change'])


temp = df

key_col = ['per_change', 'pe_ratio_ttm', 'volume_ratio'] 


temp = df#[sel_col]


##### open an existing workbood and append DataFrames as new worksheets ######
# mode='a' the workbook must exist, it will open it and append more worksheet 
# mode='w' create, overwrite if one exist, open it and write worksheets

from openpyxl import load_workbook

# create a Pandas Excel writer using openpyxl as a engine
writer = pd.ExcelWriter(f'C:\\Users\\Family_Optiplex9\\Downloads\\{export_file}.xlsx',
                        engine='openpyxl', mode='a')

dict_df = {}
for sort_col in key_col:
    df_sorted_temp = temp.sort_values(f'{sort_col}',
                                 axis=0, ascending=True,
                                 inplace=False, kind='quicksort',
                                 na_position='last', ignore_index=True,
                                 key=None)
    dict_df[f'{sort_col}'.replace(' ', '_')] = df_sorted_temp
    
    # write each dataframe to a different worksheet
    df_sorted_temp.to_excel(writer, sheet_name=f'most_active_by_{sort_col}')

# closed the Pandas Excel writer and output the Excel file
writer.save()

# =============================================================================
# for i, (sorted_by, df) in enumerate(dict_df.items()):
#     print(f'\nChecking DataFrame sorted by "{sorted_by}"')
#     print(f'{df}\n')
# 
#     if i+1 < len(dict_df.items()):
#         
#         next_df = input("\nType Enter to display next sorted DataFrame\n"
#                         "\tor anything else to break the loop:")
#         if next_df != '':
#             print('\nYou have checked all the sorted DataFrames you selected.\n')
#             break
#     
#     else:
#         print('Checked all sorted DataFrame!')
# =============================================================================
        

'''
Yahoo undervalued growth stocks screener
'''

n_per_page = 25
new_set = 0 
df = pd.DataFrame()
while new_set >= 0:
    url = 'https://finance.yahoo.com/screener/predefined/'\
        f'undervalued_growth_stocks?count={n_per_page}&offset={new_set}'
    # Pandas.read_html(): Read HTML tables into a list of DataFrame objects.
    new_df_list = pd.read_html(url)
    new_df = new_df_list[0]
    if len(new_df) == 0:
        break
    df = pd.concat([df, new_df], axis=0, ignore_index=True)
    new_set += (n_per_page + 1)
    
df.columns = df.columns.str.replace(r'%', 'per') # '%' to 'per'
df.columns = df.columns.str.replace('\W+', '_')  # whitespaces to '_'
df.columns = df.columns.str.strip('_ ')          # remove leading/trailing '_'
df.columns = map(str.lower, df.columns)          # to all lower case


# change string values to float64 for sorting
df['volume_ratio'] = str_f(df['volume']) / str_f(df['avg_vol_3_month'])   
df['per_change'] = str_f(df['per_change'])


temp = df

key_col = ['per_change', 'pe_ratio_ttm', 'volume_ratio'] 

##### open an existing workbood and append DataFrames as new worksheets ######
# mode='a' the workbook must exist, it will open it and append more worksheet 
# mode='w' create, overwrite if one exist, open it and write worksheets

from openpyxl import load_workbook

# create a Pandas Excel writer using openpyxl as a engine
writer = pd.ExcelWriter(f'C:\\Users\\Family_Optiplex9\\Downloads\\{export_file}.xlsx',
                        engine='openpyxl', mode='a')

dict_df = {}
for sort_col in key_col:
    df_sorted_temp = temp.sort_values(f'{sort_col}',
                                 axis=0, ascending=True,
                                 inplace=False, kind='quicksort',
                                 na_position='last', ignore_index=True,
                                 key=None)
    dict_df[f'{sort_col}'.replace(' ', '_')] = df_sorted_temp
    
    # write each dataframe to a different worksheet
    df_sorted_temp.to_excel(writer, sheet_name=f'undervalued_by_{sort_col}')

# closed the Pandas Excel writer and output the Excel file
writer.save()






import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import yfinance as yf

def fetch_ipo_schedule(url="https://www.nasdaq.com/market-activity/ipos"):
    """
    Scrape upcoming IPO schedule from NASDAQ website.
    Returns a list of dictionaries containing IPO details.
    """
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Find IPO table (adjust selector based on NASDAQ's structure)
        ipo_table = soup.find('table', {'class': 'market-calendar-table'})
        if not ipo_table:
            print("No IPO table found on the page.")
            return []
        
        ipo_data = []
        rows = ipo_table.find_all('tr')[1:]  # Skip header row
        
        for row in rows:
            cols = row.find_all('td')
            if len(cols) >= 5:  # Ensure enough columns
                company = cols[0].text.strip()
                ticker = cols[1].text.strip()
                exchange = cols[2].text.strip()
                price_range = cols[3].text.strip()
                ipo_date = cols[4].text.strip()
                
                # Clean price range (e.g., "$10.00 - $12.00" or "TBD")
                price_range = price_range if price_range != "TBD" else "N/A"
                
                # Validate and parse IPO date
                try:
                    ipo_date = datetime.strptime(ipo_date, "%m/%d/%Y").strftime("%Y-%m-%d")
                except ValueError:
                    ipo_date = ipo_date if ipo_date else "N/A"
                
                # Verify ticker with yfinance (check if ticker is valid)
                ticker_status = verify_ticker(ticker)
                
                ipo_data.append({
                    "Company": company,
                    "Ticker": ticker,
                    "Exchange": exchange,
                    "Price_Range": price_range,
                    "IPO_Date": ipo_date,
                    "Ticker_Valid": ticker_status
                })
        
        return ipo_data
    
    except requests.RequestException as e:
        print(f"Error fetching IPO schedule: {e}")
        return []
    except Exception as e:
        print(f"Error parsing IPO data: {e}")
        return []

def verify_ticker(ticker):
    """
    Verify if a ticker is valid using yfinance.
    Returns 'Valid' if ticker exists, 'Invalid' if not, or 'Error' if check fails.
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        if info and 'symbol' in info:
            return "Valid"
        return "Invalid"
    except Exception:
        return "Error"

def save_ipo_schedule(ipo_data, filename="upcoming_ipo_schedule.csv"):
    """
    Save IPO schedule to a CSV file.
    """
    if not ipo_data:
        print("No IPO data to save.")
        return
    
    df = pd.DataFrame(ipo_data)
    df.to_csv(filename, index=False)
    print(f"IPO schedule saved to {filename}")

def main():
    print("Fetching upcoming IPO schedule...")
    ipo_data = fetch_ipo_schedule()
    
    if ipo_data:
        print("\nUpcoming IPOs:")
        for ipo in ipo_data:
            print(f"Company: {ipo['Company']}, Ticker: {ipo['Ticker']}, "
                  f"Exchange: {ipo['Exchange']}, Price Range: {ipo['Price_Range']}, "
                  f"IPO Date: {ipo['IPO_Date']}, Ticker Valid: {ipo['Ticker_Valid']}")
        
        save_ipo_schedule(ipo_data)
    else:
        print("No upcoming IPOs found or an error occurred.")

if __name__ == "__main__":
    main()